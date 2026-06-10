"""
[2차 파싱] 다운로드된 교육과정 문서 → 학교별 개설 과목 추출
==========================================================
curriculum_files/<시도>/<학교>/ 의 '교육과정 편제/편성/배당표' 파일을 형식별로
텍스트 추출(XLSX/PDF/HWPX/HWP)한 뒤, vocab_2022.json(154개 타깃 과목)을
매칭해 학교별 개설 과목 집합을 만든다.

산출물:
  subjects_by_school.json  {idf:{school,sido,gugun,formats,n_files,subjects:{일반,진로,융합},n_subj}}
  subjects_by_school.csv   (idf, school, sido, gugun, type, subject) tidy
  parse_status.xlsx        학교별 파싱현황 (확인여부 Y / 사용파일 / 매칭과목수 / 미파싱사유)
"""
from __future__ import annotations
import sys, re, csv, json, zlib, zipfile, struct, collections
from pathlib import Path
sys.stdout.reconfigure(encoding="utf-8")

import fitz  # PyMuPDF
from openpyxl import load_workbook

try:
    fitz.TOOLS.mupdf_display_errors(False)  # 무해한 주석 경고 억제
except Exception:
    pass

ROOT = Path(__file__).parent
FILES_DIR = ROOT / "curriculum_files"
VOCAB = json.loads((ROOT / "vocab_2022.json").read_text(encoding="utf-8"))
OUT_JSON = ROOT / "subjects_by_school.json"
OUT_CSV = ROOT / "subjects_by_school.csv"
STATUS_XLSX = ROOT / "parse_status.xlsx"

CURRI_KW = ("배당", "편제", "편성", "교육과정")
JUNK_KW = ("학사일정", "일정표", "체험활동", "수학여행", "현장체험",
           "학교평가", "평가 결과", "수업시수", "시정표")


# ── 텍스트 정규화 & 매칭 ────────────────────────────────────
def normspace(s):
    """로마숫자 통일 + 공백 단일화(매칭 경계 보존)."""
    s = s.replace("Ⅰ", "I").replace("Ⅱ", "II").replace("Ⅲ", "III").replace("Ⅳ", "IV")
    return re.sub(r"\s+", " ", s)


def _compile(term):
    """과목명 → 경계+유연공백 정규식. 한글/영문 안에 묻힌 부분문자열 오탐 방지."""
    t = normspace(term).strip()
    parts = [re.escape(p) for p in t.split(" ") if p]
    body = r"\s*".join(parts)               # 공백은 0개 이상 허용
    return re.compile(r"(?<![가-힣A-Za-z])" + body + r"(?![가-힣A-Za-z])")


# 중복 nkey 제거(VOCAB는 빈도순 → 첫 항목 채택), 긴 이름 우선
_seen = {}
for _k, _v in VOCAB.items():
    _nk = re.sub(r"\s+", "", normspace(_k))
    if _nk and _nk not in _seen:
        _seen[_nk] = (_k, _v["type"])
_VOCAB_N = sorted(((name, typ, _compile(name)) for name, typ in _seen.values()),
                  key=lambda x: -len(x[0]))


def match_subjects(text):
    t = normspace(text)
    found = {}
    for name, typ, rx in _VOCAB_N:
        if rx.search(t):
            found[name] = typ
    return found


# ── 형식별 텍스트 추출 ──────────────────────────────────────
SKIP_SHEET = ("요령", "작성", "방법", "기준", "코드", "안내", "점검", "피드백",
              "메뉴", "db", "list", "삭제", "참고", "예시", "과목기본정보")
REF_SHEET_MAX = 100   # 한 시트가 이 이상 매칭하면 전체 과목 참조목록으로 보고 제외


def match_xlsx(path):
    """시트별로 매칭 후, 참조목록 시트(과다매칭)는 빼고 union."""
    wb = load_workbook(path, data_only=True, read_only=True)
    found = {}
    for ws in wb.worksheets:
        if any(k in ws.title.lower() for k in SKIP_SHEET):
            continue
        txt = " ".join(c for row in ws.iter_rows(values_only=True) for c in row
                       if isinstance(c, str) and c.strip())
        hits = match_subjects(txt)
        if 0 < len(hits) < REF_SHEET_MAX:
            found.update(hits)
    wb.close()
    return found


def text_pdf(path):
    out = []
    with fitz.open(path) as doc:
        for pg in doc:
            out.append(pg.get_text())
    return " ".join(out)


def text_hwpx(path):
    out = []
    with zipfile.ZipFile(path) as z:
        for n in z.namelist():
            if n.startswith("Contents/") and n.endswith(".xml"):
                xml = z.read(n).decode("utf-8", "ignore")
                out.append(re.sub(r"<[^>]+>", " ", xml))
    return " ".join(out)


def text_hwp(path):
    """HWP5(OLE) BodyText 섹션 zlib 해제 후 PARA_TEXT(tag 0x43) 추출."""
    import olefile
    if not olefile.isOleFile(str(path)):
        return ""
    ole = olefile.OleFileIO(str(path))
    try:
        header = ole.openstream("FileHeader").read()
        compressed = bool(header[36] & 1)
        texts = []
        for entry in ole.listdir():
            if entry[0] == "BodyText":
                data = ole.openstream(entry).read()
                if compressed:
                    try:
                        data = zlib.decompress(data, -15)
                    except Exception:
                        continue
                texts.append(_hwp_records_text(data))
        return " ".join(texts)
    finally:
        ole.close()


def _hwp_records_text(buf):
    out = []
    i, n = 0, len(buf)
    while i + 4 <= n:
        hdr = struct.unpack_from("<I", buf, i)[0]
        i += 4
        tag = hdr & 0x3FF
        size = (hdr >> 20) & 0xFFF
        if size == 0xFFF:
            if i + 4 > n:
                break
            size = struct.unpack_from("<I", buf, i)[0]
            i += 4
        if i + size > n:
            break
        if tag == 0x43:  # HWPTAG_PARA_TEXT (UTF-16LE, with inline ctrl chars)
            raw = buf[i:i + size]
            try:
                s = raw.decode("utf-16-le", "ignore")
            except Exception:
                s = ""
            s = "".join(ch for ch in s if ch == " " or ord(ch) >= 0x20)
            out.append(s)
        i += size
    return " ".join(out)


# 텍스트 추출기(blob 반환). xlsx는 시트인지 매칭이 필요해 별도 처리.
EXTRACTORS = {".pdf": text_pdf, ".hwpx": text_hwpx, ".hwp": text_hwp}
XLSX_EXT = {".xlsx", ".xlsm"}


def subjects_in_file(path):
    """파일 1개 → {과목:유형}. 형식별 분기."""
    ext = path.suffix.lower()
    if ext in XLSX_EXT:
        return match_xlsx(path)
    fn = EXTRACTORS.get(ext)
    if not fn:
        return None  # 미지원
    return match_subjects(fn(path))


def is_curriculum(fn):
    return any(k in fn for k in CURRI_KW) and not any(k in fn for k in JUNK_KW)


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--include-all", action="store_true",
                    help="교육과정 키워드 외 파일도 폴백 파싱")
    args = ap.parse_args()

    # (시도,학교명) -> shl_idf_cd, gugun  매핑 (동명 학교 충돌 방지)
    idmap = {}
    with open(ROOT / "school_list.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            idmap[(r["sido"], r["school_name"])] = (r["shl_idf_cd"], r["gugun"])

    schools = [d for sido in sorted(FILES_DIR.iterdir()) if sido.is_dir()
               for d in sorted(sido.iterdir()) if d.is_dir()]
    if args.limit:
        schools = schools[:args.limit]
    print(f"학교 폴더 {len(schools)}개 파싱")

    big = {}
    status_rows = []
    tidy = open(OUT_CSV, "w", newline="", encoding="utf-8-sig")
    tw = csv.writer(tidy)
    tw.writerow(["shl_idf_cd", "school", "sido", "gugun", "type", "subject"])

    for k, d in enumerate(schools, 1):
        sido, name = d.parent.name, d.name
        idf, gugun = idmap.get((sido, name), ("", ""))
        files = list(d.iterdir())
        cur = [f for f in files if is_curriculum(f.name)]
        if not cur and args.include_all:
            cur = [f for f in files if f.suffix.lower() in EXTRACTORS]
        used, fmts, subjects, reasons = [], set(), {}, []
        for f in cur:
            ext = f.suffix.lower()
            try:
                hits = subjects_in_file(f)
            except Exception as e:
                reasons.append(f"err:{ext}:{str(e)[:25]}")
                continue
            if hits is None:
                reasons.append(f"noext:{ext}")
                continue
            if hits:
                used.append(f.name)
                fmts.add(ext)
                subjects.update(hits)
        bytype = collections.defaultdict(list)
        for s, t in subjects.items():
            bytype[t].append(s)
            tw.writerow([idf, name, sido, gugun, t, s])
        if subjects:
            big[idf or name] = {"shl_idf_cd": idf, "school": name, "sido": sido,
                                "gugun": gugun, "formats": sorted(fmts),
                                "n_files": len(used), "used_files": used,
                                "subjects": {t: sorted(v) for t, v in bytype.items()},
                                "n_subj": len(subjects)}
        status_rows.append({
            "idx": k, "shl_idf_cd": idf, "school": name, "sido": sido,
            "gugun": gugun, "확인여부": "Y",
            "후보파일수": len(cur), "사용파일수": len(used),
            "매칭과목수": len(subjects),
            "유형분포": " ".join(f"{t}{len(v)}" for t, v in bytype.items()),
            "사용파일": " | ".join(used),
            "미파싱사유": " ".join(reasons) if not subjects else "",
        })
        if k % 100 == 0 or args.limit:
            print(f"[{k}/{len(schools)}] {name} files={len(cur)} subj={len(subjects)}")
    tidy.close()
    OUT_JSON.write_text(json.dumps(big, ensure_ascii=False, indent=1), encoding="utf-8")
    write_status_xlsx(status_rows)
    n_ok = sum(1 for r in status_rows if r["매칭과목수"] > 0)
    print(f"\n완료: {len(schools)}교 중 과목추출 성공 {n_ok}교 -> {OUT_JSON.name}")


def write_status_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "파싱현황"
    cols = ["idx", "shl_idf_cd", "school", "sido", "gugun", "확인여부",
            "후보파일수", "사용파일수", "매칭과목수", "유형분포",
            "사용파일", "미파싱사유"]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    for i, wdt in enumerate([6, 38, 30, 14, 10, 8, 9, 9, 9, 22, 50, 40], 1):
        ws.column_dimensions[chr(64 + i)].width = wdt
    ws.freeze_panes = "A2"
    wb.save(STATUS_XLSX)


if __name__ == "__main__":
    main()
