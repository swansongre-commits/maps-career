"""
학교알리미 전국 고교 '교육과정 편성(항목05)' 첨부 일괄 다운로더  [1차 수집]
==========================================================================
school_list.csv 의 모든 고교를 순회하며, 각 학교가 '학교교육과정 편성·운영·평가'
(공시항목 JG_HANGMOK_CD=05) 아래 올린 첨부파일을 형식 무관(XLSX/PDF/HWP/HWPX)
으로 모두 내려받는다. 파싱은 2차 작업에서 따로 한다.

검증된 사실(2026-05-29):
  - 학교검색 getSchoolList.do : Python은 WAF에 막힘 → school_list.csv 시드 사용
    (필요시 --refresh-list 가 curl로 갱신).
  - 다운로드 EiFileDownLoad.do : Python requests로 정상. JG_HANGMOK_CD=05,
    JG_YEAR(공시연도) 아래 FILE_SEQ를 1..N 열거. 실제 첨부면 octet-stream +
    Content-Disposition(파일명), 없으면 302. 중간 빈 SEQ가 있을 수 있어 끝까지 탐색.
  - 첨부 형식은 학교마다 제각각(표본 30교: PDF22·HWP8·HWPX7·XLSX7, 첨부없음 7).

산출물:
  curriculum_files/<시도>/<학교명>/<원본파일>        다운로드 원본
  collection_status.csv   체크포인트(재개용, 실시간 기록, utf-8-sig)
  collection_status.xlsx  최종 수집현황 엑셀 (확인여부 Y / 파일수 / 교육과정파일수)

사용:
  python crawl_curriculum.py --year 2025                  # 전국
  python crawl_curriculum.py --year 2025 --sido 서울특별시
  python crawl_curriculum.py --year 2025 --limit 20       # 테스트
  python crawl_curriculum.py --refresh-list               # 목록 갱신(curl)
  python crawl_curriculum.py --build-xlsx                 # CSV→엑셀만 재생성
"""
from __future__ import annotations
import re
import csv
import sys
import json
import time
import argparse
import subprocess
from pathlib import Path
from urllib.parse import unquote

import requests

BASE = "https://www.schoolinfo.go.kr"
ROOT = Path(__file__).parent
LIST_CSV = ROOT / "school_list.csv"
FILES_DIR = ROOT / "curriculum_files"
STATUS_CSV = ROOT / "collection_status.csv"
STATUS_XLSX = ROOT / "collection_status.xlsx"
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

# 교육과정 편제/편성/배당표로 보이는 파일 판별(2차 파싱 대상 후보)
CURRI_KW = ("배당", "편제", "편성", "교육과정")
JUNK_KW = ("학사일정", "체험활동", "수학여행", "현장체험", "학교평가", "평가 결과")
STATUS_FIELDS = ["idx", "shl_idf_cd", "school_name", "sido", "gugun", "crse_cd",
                 "확인여부", "상태", "총파일수", "교육과정파일수", "파일목록", "저장폴더"]


def sanitize(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", s or "").strip() or "_"


# ── 학교목록 (시드 CSV / curl 갱신) ─────────────────────────
def refresh_school_list(word="고등학교"):
    url = f"{BASE}/ei/ss/pneiss_a04_s0/getSchoolList.do"
    p = subprocess.run(["curl", "-s", "--data", f"SEARCH_WORD={word}", url],
                       capture_output=True, text=True, encoding="utf-8")
    data = json.loads(p.stdout)
    cols = ["SHL_NM", "SHL_IDF_CD", "SHL_CD", "SHL_CRSE_SC_CD",
            "USER_DFN_CODE_VALUE_01", "USER_DFN_CODE_VALUE_02",
            "SIDO_CODE", "GUGUN_CODE", "FOND_SC_NM", "FULL_ADDR"]
    with open(LIST_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["school_name", "shl_idf_cd", "shl_cd", "crse_cd", "sido",
                    "gugun", "sido_code", "gugun_code", "found", "addr"])
        for x in data:
            w.writerow([x.get(c, "") for c in cols])
    print(f"[list] {len(data)}개교 저장 -> {LIST_CSV}")
    return len(data)


def load_school_list(sido=None, crse="04"):
    rows = []
    with open(LIST_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if crse and r["crse_cd"] != crse:
                continue
            if sido and sido not in r["sido"]:
                continue
            rows.append(r)
    return rows


# ── 항목05 첨부 다운로드 (FILE_SEQ 열거) ────────────────────
def download_item05(session, idf, year, dest_dir, max_seq=12, retries=3, pause=0.25):
    """JG_HANGMOK_CD=05 아래 FILE_SEQ를 끝까지 열거해 모든 첨부를 저장.
    반환: [(filename, ext, bytes, is_curriculum)]"""
    saved = []
    for seq in range(0, max_seq + 1):   # FILE_SEQ는 0부터 시작(첫 첨부가 0번인 학교 다수)
        params = {"SHL_IDF_CD": idf, "JG_BURYU_CD": "JG020", "JG_HANGMOK_CD": "05",
                  "JG_GUBUN": "1", "JG_YEAR": str(year), "JG_CHASU": "1",
                  "USE_YN": "Y", "FILE_SEQ": str(seq)}
        r = None
        for attempt in range(retries):
            try:
                r = session.get(f"{BASE}/servlets/EiFileDownLoad.do", params=params,
                                timeout=60, allow_redirects=False)
                break
            except requests.RequestException:
                time.sleep(1 + attempt)
        if r is None or "octet-stream" not in r.headers.get("Content-Type", ""):
            continue  # 이 SEQ엔 파일 없음(302) → 중간 빈번호 가능, 계속
        cd = r.headers.get("Content-Disposition", "")
        m = re.search(r"filename=(.+?);?$", cd)
        fn = unquote(m.group(1)).strip() if m else f"{idf}_{year}_{seq}"
        # 첨부가 아닌 응답(점검/에러 HTML) 방어: 알려진 시그니처만 저장
        sig = r.content[:4]
        if not (sig[:2] == b"PK" or sig == b"\xd0\xcf\x11\xe0" or sig == b"%PDF"
                or fn.lower().endswith((".hwp", ".hwpx", ".pdf", ".xlsx", ".xls",
                                        ".doc", ".docx"))):
            continue
        dest_dir.mkdir(parents=True, exist_ok=True)
        path = dest_dir / sanitize(fn)
        path.write_bytes(r.content)
        ext = path.suffix.lower()
        is_cur = any(k in fn for k in CURRI_KW) and not any(k in fn for k in JUNK_KW)
        saved.append((fn, ext, len(r.content), is_cur))
        time.sleep(pause)
    return saved


# ── 체크포인트/엑셀 ─────────────────────────────────────────
def load_done():
    done = set()
    if STATUS_CSV.exists():
        with open(STATUS_CSV, encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                if r.get("확인여부") == "Y":
                    done.add(r["shl_idf_cd"])
    return done


def build_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "수집현황"
    ws.append(STATUS_FIELDS)
    head = ws[1]
    for c in head:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center")
    n = ok = 0
    if STATUS_CSV.exists():
        with open(STATUS_CSV, encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                ws.append([r.get(k, "") for k in STATUS_FIELDS]); n += 1
                if r.get("상태") == "ok":
                    ok += 1
    widths = [6, 38, 22, 12, 8, 8, 8, 10, 9, 12, 60, 30]
    for i, wcol in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = wcol
    ws.freeze_panes = "A2"
    wb.save(STATUS_XLSX)
    print(f"[xlsx] {n}개교 기록(파일있음 {ok}) -> {STATUS_XLSX}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", type=int, default=2025)
    ap.add_argument("--sido", default=None)
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--pause", type=float, default=0.25, help="요청간 대기(초)")
    ap.add_argument("--max-seq", type=int, default=12)
    ap.add_argument("--refresh-list", action="store_true")
    ap.add_argument("--build-xlsx", action="store_true", help="CSV→엑셀만 재생성")
    args = ap.parse_args()

    if args.build_xlsx:
        build_xlsx(); return
    if args.refresh_list or not LIST_CSV.exists():
        refresh_school_list()

    schools = load_school_list(sido=args.sido)
    if args.limit:
        schools = schools[:args.limit]
    done = load_done()
    todo = [s for s in schools if s["shl_idf_cd"] not in done]
    print(f"대상 {len(schools)}개교 / 미완료 {len(todo)}개 (완료 {len(done)} 건너뜀), "
          f"year={args.year}")

    session = requests.Session(); session.headers.update(HEADERS)
    new = not STATUS_CSV.exists()
    fh = open(STATUS_CSV, "a", newline="", encoding="utf-8-sig")
    w = csv.DictWriter(fh, fieldnames=STATUS_FIELDS)
    if new:
        w.writeheader()

    base_idx = len(done)
    for i, sc in enumerate(todo, 1):
        idf, name = sc["shl_idf_cd"], sc["school_name"]
        sido, gugun = sc["sido"], sc["gugun"]
        dest = FILES_DIR / sanitize(sido) / sanitize(name)
        try:
            saved = download_item05(session, idf, args.year, dest,
                                    max_seq=args.max_seq, pause=args.pause)
            status = "ok" if saved else "no_file"
        except Exception as e:
            saved, status = [], f"error:{str(e)[:40]}"
        n_cur = sum(1 for _, _, _, c in saved if c)
        w.writerow({
            "idx": base_idx + i, "shl_idf_cd": idf, "school_name": name,
            "sido": sido, "gugun": gugun, "crse_cd": sc["crse_cd"],
            "확인여부": "Y", "상태": status, "총파일수": len(saved),
            "교육과정파일수": n_cur,
            "파일목록": " | ".join(fn for fn, *_ in saved),
            "저장폴더": str(dest) if saved else "",
        })
        fh.flush()
        if i % 20 == 0 or args.limit:
            print(f"[{i}/{len(todo)}] {name} -> {status} "
                  f"files={len(saved)} curri={n_cur}")
    fh.close()
    build_xlsx()
    print("완료.")


if __name__ == "__main__":
    main()
