"""
학교알리미 → 고교별 '교육과정 학점배당표(2022 개정)' XLSX 수집·과목 추출

핵심 발견(2026-05-29 라이브 검증):
  - 공시상세 목록페이지 Pneipp_b14_s0p.do 는 연간 공시갱신 점검중 차단됨.
  - 그러나 파일 다운로드 EiFileDownLoad.do 는 점검과 무관하게 동작.
    => 목록을 안 거치고 FILE_SEQ 를 1,2,3... 직접 열거하면 된다.
       실제 파일이면 Content-Type=application/octet-stream + Content-Disposition(파일명),
       없으면 302 리다이렉트(HTML). 파일명에 'YYYY학년도 입학생' 이 들어있어
       입학년도/개정연도를 그대로 알 수 있다.

흐름:
  [1] 학교검색  POST getSchoolList.do (JSON)      ✅ 점검중에도 동작
  [2] 파일다운  GET  EiFileDownLoad.do (XLSX)     ✅ 점검중에도 동작 (FILE_SEQ 열거)
  [3] XLSX 파싱 openpyxl: 과목유형/과목 컬럼 추출  ✅ 양재고로 검증(90과목)
"""
import re
import sys
import json
import requests
from pathlib import Path
from urllib.parse import unquote
from openpyxl import load_workbook

BASE = "https://www.schoolinfo.go.kr"
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
           "Referer": f"{BASE}/Main.do", "X-Requested-With": "XMLHttpRequest"}
OUT = Path(__file__).parent / "curriculum_files"


def new_session():
    s = requests.Session(); s.headers.update(HEADERS)
    s.get(f"{BASE}/Main.do", timeout=20)
    return s


# ── [1] 학교검색 (SHL_CRSE_SC_CD 04 = 고등학교) ──────────────
def search_school(s, word, gugun=None, crse="04"):
    r = s.post(f"{BASE}/ei/ss/pneiss_a04_s0/getSchoolList.do",
               data={"SEARCH_WORD": word}, timeout=30)
    out = []
    for x in r.json():
        if crse and x.get("SHL_CRSE_SC_CD") != crse:
            continue
        if gugun and gugun not in (x.get("USER_DFN_CODE_VALUE_02") or ""):
            continue
        out.append(x)
    return out


# ── [2] 학점배당표 첨부 직접 다운로드 (FILE_SEQ 열거) ────────
def fetch_curriculum_files(s, shl_idf_cd, year, max_seq=10):
    """JG_YEAR(공시연도) 아래 FILE_SEQ 를 열거해 실제 첨부를 모두 받는다.
    반환: [{seq, filename, ext, intake_year, revision, path}]"""
    OUT.mkdir(exist_ok=True)
    found = []
    for seq in range(1, max_seq + 1):
        params = {"SHL_IDF_CD": shl_idf_cd, "JG_BURYU_CD": "JG020",
                  "JG_HANGMOK_CD": "05", "JG_GUBUN": "1", "JG_YEAR": str(year),
                  "JG_CHASU": "1", "USE_YN": "Y", "FILE_SEQ": str(seq)}
        r = s.get(f"{BASE}/servlets/EiFileDownLoad.do", params=params,
                  timeout=60, allow_redirects=False)
        ctype = r.headers.get("Content-Type", "")
        if "octet-stream" not in ctype:      # 302 등 → 파일 없음, 열거 종료
            break
        cd = r.headers.get("Content-Disposition", "")
        m = re.search(r"filename=(.+?);?$", cd)
        fn = unquote(m.group(1)).strip() if m else f"{shl_idf_cd}_{year}_{seq}.bin"
        path = OUT / fn
        path.write_bytes(r.content)
        ymatch = re.search(r"(\d{4})\s*학년도", fn)
        intake = int(ymatch.group(1)) if ymatch else None
        rev = "2022" if "2022" in fn else ("2015" if "2015" in fn else None)
        if rev is None and intake:           # 파일명에 개정연도 없으면 입학년도로 추정
            rev = "2022" if intake >= 2025 else "2015"
        found.append({"seq": seq, "filename": fn, "ext": path.suffix.lower(),
                      "intake_year": intake, "revision": rev, "path": str(path)})
    return found


# ── [3] XLSX → 과목 (유형별) ─────────────────────────────────
def parse_subjects(xlsx_path, type_col=3, subj_col=4, data_start=5):
    """학점배당표 시트에서 (과목유형, 과목명) 추출.
    양재고 양식 기준 col3=과목유형(공통/일반/융합/진로), col4=과목.
    학교별 양식 차이가 있으면 헤더 자동탐지로 보정."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = next((wb[n] for n in wb.sheetnames if "입학" in n), wb.worksheets[0])
    # 헤더 자동탐지(과목유형/과목 위치)
    for r in range(1, 8):
        for j in range(1, ws.max_column + 1):
            v = str(ws.cell(r, j).value or "").replace(" ", "")
            if v.endswith("과목유형"):
                type_col = j
            elif re.search(r"\)?과목$", v) and "유형" not in v:
                subj_col = j
    by_type = {}
    for r in range(data_start, ws.max_row + 1):
        sub = ws.cell(r, subj_col).value
        if not sub or not str(sub).strip():
            continue
        typ = str(ws.cell(r, type_col).value or "").strip()
        by_type.setdefault(typ, []).append(str(sub).strip())
    return by_type


if __name__ == "__main__":
    s = new_session()
    hits = search_school(s, "양재", gugun="서초구")
    if not hits:
        sys.exit("학교 검색 결과 없음")
    school = hits[0]
    print(f"{school['SHL_NM']}  SHL_IDF_CD={school['SHL_IDF_CD']}")
    files = fetch_curriculum_files(s, school["SHL_IDF_CD"], year=2025)
    print(json.dumps(files, ensure_ascii=False, indent=2))
    for f in files:
        if f["ext"] == ".xlsx":
            print(f"\n=== {f['filename']} ===")
            for typ, subs in parse_subjects(f["path"]).items():
                print(f"[{typ}] {len(subs)}개: {', '.join(subs)}")
